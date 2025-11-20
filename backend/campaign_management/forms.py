# campaign_management/forms.py

from django import forms
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.db import models
from django.db.models import Q
from datetime import datetime, time
from PIL import Image  # ensure Pillow is installed
import io
import os
import uuid
import re

from .models import Campaign, CampaignAssignment, CampaignCollateral
from user_management.models import User  # your custom User model

DATE_FMT = "%d/%m/%Y"  # dd/mm/yyyy
REQUIRED_PRINT_COLUMNS = {
    "Collateral Name",
    "Schedule Date of Collaterals",
    "Delivery Date of Collateral",
    "Delivery Address",
    "Contact Person Name",
    "Contact Person Phone Number",
}

# Campaign Creation Form
class CampaignForm(forms.ModelForm):
    # Add brand_campaign_id field
    brand_campaign_id = forms.CharField(
        label="Brand–Campaign ID",
        required=False,
        help_text="Leave blank to auto-generate (e.g., BRAND-ABC123)",
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Leave blank to auto-generate'})
    )
    
    start_date = forms.DateField(
        label="Start Date of the Campaign",
        widget=forms.DateInput(attrs={"type": "date", "class": "form-control"})
    )
    end_date = forms.DateField(
        label="End Date of the Campaign",
        widget=forms.DateInput(attrs={"type": "date", "class": "form-control"})
    )

    # Printing toggle
    printing_required = forms.ChoiceField(
        choices=(("no", "No"), ("yes", "Yes")),
        label="Does this campaign require printing of collateral by Inditech?",
        widget=forms.Select(attrs={"class": "form-select"})
    )
    printing_excel = forms.FileField(
        required=False,
        label="Upload Excel (required if 'Yes')",
        help_text="XLSX with columns: " + ", ".join(sorted(REQUIRED_PRINT_COLUMNS)),
    )

    class Meta:
        model = Campaign
        fields = [
            'brand_campaign_id',  # Added as first field
            'name',
            'brand_name',
            'company_name',
            'incharge_name',
            'incharge_contact',
            'incharge_designation',
            'num_doctors',
            'items_per_clinic_per_year',
            'start_date',
            'end_date',
            'contract',
            'brand_logo',
            'company_logo',
            'printing_required',
            'printing_excel',
            'description',
            'status',
        ]
        widgets = {
            'name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Campaign Name'}),
            'brand_name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Brand Name'}),
            'company_name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Company Name'}),
            'incharge_name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Incharge Name'}),
            'incharge_contact': forms.TextInput(attrs={'class': 'form-control', 'placeholder': '+91XXXXXXXXXX or 10 digits'}),
            'incharge_designation': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Designation'}),
            'num_doctors': forms.NumberInput(attrs={'class': 'form-control', 'min': 0, 'placeholder': 'Number of Doctors'}),
            'items_per_clinic_per_year': forms.NumberInput(attrs={'class': 'form-control', 'min': 0, 'placeholder': 'Items per Clinic per Year'}),
            'contract': forms.ClearableFileInput(attrs={'class': 'form-control', 'accept': '.pdf,.doc,.docx'}),
            'brand_logo': forms.ClearableFileInput(attrs={'class': 'form-control', 'accept': 'image/*'}),
            'company_logo': forms.ClearableFileInput(attrs={'class': 'form-control', 'accept': 'image/*'}),
            'description': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Campaign Description'}),
            'status': forms.Select(attrs={'class': 'form-select'}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # When editing, prefill date fields and normalize toggles
        if self.instance and getattr(self.instance, 'pk', None):
            try:
                if getattr(self.instance, 'start_date', None):
                    self.fields['start_date'].initial = timezone.localtime(self.instance.start_date).date()
                if getattr(self.instance, 'end_date', None):
                    self.fields['end_date'].initial = timezone.localtime(self.instance.end_date).date()
            except Exception:
                if getattr(self.instance, 'start_date', None):
                    self.fields['start_date'].initial = getattr(self.instance.start_date, 'date', lambda: None)() or self.instance.start_date.date()
                if getattr(self.instance, 'end_date', None):
                    self.fields['end_date'].initial = getattr(self.instance.end_date, 'date', lambda: None)() or self.instance.end_date.date()

            self.fields['printing_required'].initial = 'yes' if getattr(self.instance, 'printing_required', False) else 'no'
            # Ensure brand_campaign_id shows existing value
            if hasattr(self.instance, 'brand_campaign_id'):
                self.fields['brand_campaign_id'].initial = (self.instance.brand_campaign_id or '')

    # —— helpers ——
    def _parse_date(self, value) -> datetime:
        if value is None:
            raise ValueError("Date cannot be empty")
        if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day') and not isinstance(value, str):
            dt_date = value
        else:
            s = str(value).strip()
            if not s:
                raise ValueError("Date cannot be empty")
            try:
                dt = datetime.strptime(s, "%Y-%m-%d")
            except Exception:
                dt = datetime.strptime(s, DATE_FMT)
            dt_date = dt.date()
        return timezone.make_aware(datetime.combine(dt_date, time.min))

    def _parse_date_end(self, value) -> datetime:
        if value is None:
            raise ValueError("Date cannot be empty")
        if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day') and not isinstance(value, str):
            dt_date = value
        else:
            s = str(value).strip()
            if not s:
                raise ValueError("Date cannot be empty")
            try:
                dt = datetime.strptime(s, "%Y-%m-%d")
            except Exception:
                dt = datetime.strptime(s, DATE_FMT)
            dt_date = dt.date()
        return timezone.make_aware(datetime.combine(dt_date, time.max.replace(microsecond=0)))

    def _validate_logo(self, f, which: str):
        if not f:
            return
        try:
            # Read into Pillow without saving
            img = Image.open(f)
            w, h = img.size
            if (w, h) != (200, 70):
                raise ValidationError(f"{which} must be exactly 200×70 pixels (got {w}×{h}).")
            
            # Validate file size (max 2MB)
            if hasattr(f, 'size') and f.size > 2 * 1024 * 1024:
                raise ValidationError(f"{which} must be less than 2MB.")
                
        except Exception as e:
            if isinstance(e, ValidationError):
                raise e
            raise ValidationError(f"Invalid image file for {which}.")
        finally:
            # Reset file pointer for Django to save later
            if hasattr(f, 'seek'):
                f.seek(0)

    def _validate_printing_excel(self, f):
        if not f:
            return
            
        name = getattr(f, 'name', '') or ''
        ext = os.path.splitext(name)[1].lower()
        if ext != '.xlsx':
            raise ValidationError("Printing plan must be an Excel .xlsx file.")
            
        # Validate file size (max 10MB)
        if hasattr(f, 'size') and f.size > 10 * 1024 * 1024:
            raise ValidationError("Excel file must be less than 10MB.")
        
        # Try to validate headers via openpyxl (optional dependency)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            
            # Check if worksheet has data
            if ws.max_row == 0:
                raise ValidationError("Excel file is empty.")
                
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            missing = REQUIRED_PRINT_COLUMNS - set(headers)
            if missing:
                raise ValidationError(f"Printing plan is missing columns: {', '.join(sorted(missing))}")
                
        except ImportError:
            # If openpyxl is not installed, skip header validation (server-side) and accept the file.
            pass
        except Exception as e:
            if isinstance(e, ValidationError):
                raise e
            raise ValidationError("Error reading Excel file. Please check the format.")
        finally:
            if hasattr(f, 'seek'):
                f.seek(0)

    # —— main clean ——
    def clean(self):
        data = super().clean()
        
        # Dates validation
        start_date_val = self.cleaned_data.get('start_date', None)
        end_date_val = self.cleaned_data.get('end_date', None)
        
        # Dates are required in the model, so they must be provided
        if not start_date_val:
            raise ValidationError("Start Date is required.")
        if not end_date_val:
            raise ValidationError("End Date is required.")
        
        try:
            data['start_date'] = self._parse_date(start_date_val)
            data['end_date'] = self._parse_date_end(end_date_val)
        except ValueError as e:
            raise ValidationError("Enter valid Start/End dates.")
        except Exception:
            raise ValidationError("Enter valid Start/End dates.")
            
        if data['end_date'] < data['start_date']:
            raise ValidationError("End Date cannot be earlier than Start Date.")

        # Normalize incharge_contact (allow 10-digit or +91XXXXXXXXXX) - make optional
        raw_contact = (data.get('incharge_contact') or '').strip()
        if raw_contact:
            digits = ''.join(ch for ch in raw_contact if ch.isdigit())
            if len(digits) == 10:
                data['incharge_contact'] = '+91' + digits
            elif digits.startswith('91') and len(digits) == 12:
                data['incharge_contact'] = '+' + digits
            elif len(digits) == 12 and digits.startswith('91'):
                data['incharge_contact'] = '+' + digits
            else:
                raise ValidationError("Enter a valid 10-digit phone number or +91XXXXXXXXXX format.")
        else:
            # If no contact provided, set to empty string (since field is optional in model)
            data['incharge_contact'] = ''

        # Logos dimension validation - make optional
        brand_logo = self.cleaned_data.get('brand_logo')
        company_logo = self.cleaned_data.get('company_logo')
        
        # Only validate if logos are actually uploaded
        if brand_logo and hasattr(brand_logo, 'name') and brand_logo.name:
            self._validate_logo(brand_logo, "Brand Logo")
        if company_logo and hasattr(company_logo, 'name') and company_logo.name:
            self._validate_logo(company_logo, "Company Logo")

        # Printing Excel when required - FIXED: Normalize printing_required to boolean
        pr = self.cleaned_data.get('printing_required')
        f = self.cleaned_data.get('printing_excel')
        
        # Convert choice field to boolean
        if isinstance(pr, str):
            data['printing_required'] = (pr.lower() == 'yes')
        else:
            data['printing_required'] = bool(pr)

        # Consider existing uploaded file when editing
        existing_file = getattr(self.instance, 'printing_excel', None)
        if data['printing_required']:
            if not f and not existing_file:
                raise ValidationError("Upload the Excel file because 'Printing required' is Yes.")
            if f:
                self._validate_printing_excel(f)
        else:
            # If printing not required, clear any uploaded file from this submission
            if f:
                data['printing_excel'] = None
            # If no new file uploaded but we have existing file, preserve it
            elif existing_file:
                data['printing_excel'] = existing_file

        # Auto-generate brand_campaign_id if not supplied
        bcid = (self.cleaned_data.get('brand_campaign_id') or '').strip()
        if not bcid:
            base = (self.cleaned_data.get('brand_name') or self.cleaned_data.get('name') or 'CMP')
            # Clean and format the base string
            base = re.sub(r'[^A-Za-z0-9]+', '-', base).strip('-').upper()[:12]
            if not base:
                base = 'CMP'
            data['brand_campaign_id'] = f"{base}-{uuid.uuid4().hex[:6].upper()}"
        else:
            # Validate provided brand_campaign_id format
            if not re.match(r'^[A-Za-z0-9\-_]+$', bcid):
                raise ValidationError("Brand Campaign ID can only contain letters, numbers, hyphens, and underscores.")
            data['brand_campaign_id'] = bcid.strip()

        # Validate required fields (keep only core identifiers required)
        required_fields = ['name', 'brand_name']
        for field in required_fields:
            if not data.get(field):
                raise ValidationError(f"{field.replace('_', ' ').title()} is required.")

        return data

    def save(self, commit=True):
        # Ensure printing_required is properly set before saving
        instance = super().save(commit=False)
        instance.printing_required = self.cleaned_data.get('printing_required', False)
        
        if commit:
            instance.save()
            self.save_m2m()
        return instance


# Field Rep Assignment Form
class CampaignAssignmentForm(forms.ModelForm):
    field_rep = forms.ModelChoiceField(
        queryset=User.objects.filter(role='field_rep', active=True),
        label="Select Field Rep",
        widget=forms.Select(attrs={'class': 'form-select'})
    )

    class Meta:
        model = CampaignAssignment
        fields = ['campaign', 'field_rep']
        widgets = {
            'campaign': forms.Select(attrs={'class': 'form-select'}),
        }

    def clean(self):
        cleaned_data = super().clean()
        campaign = cleaned_data.get('campaign')
        field_rep = cleaned_data.get('field_rep')

        if campaign and field_rep:
            # Check if this field rep is already assigned to this campaign
            existing_assignment = CampaignAssignment.objects.filter(
                campaign=campaign,
                field_rep=field_rep
            ).exists()
            
            if existing_assignment and (not self.instance or not self.instance.pk):
                raise ValidationError("This field rep is already assigned to the selected campaign.")

        return cleaned_data


from collateral_management.models import Collateral as CMCollateral
from collateral_management.models import CampaignCollateral as CMCampaignCollateral
# Collateral Assignment Form
class CampaignCollateralForm(forms.ModelForm):
    campaign = forms.CharField(
        max_length=255,
        required=False,
        widget=forms.TextInput(attrs={
            'class': 'form-control', 
            'readonly': 'readonly',
            'style': 'background-color: #e9ecef; cursor: not-allowed;'
        }),
        label="Brand Campaign ID",
        disabled=True
    )
    collateral = forms.ModelChoiceField(
        queryset=CMCollateral.objects.none(),  # Will be set in __init__
        label="Select Collateral",
        widget=forms.Select(attrs={'class': 'form-select'})
    )
    
    class Meta:
        # Use the bridging model from collateral_management to align with Collateral
        model = CMCampaignCollateral
        fields = ['collateral', 'start_date', 'end_date']
        widgets = {
            'start_date': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'end_date': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
        }
    
    def __init__(self, *args, **kwargs):
        brand_campaign_id = kwargs.pop('brand_campaign_id', None)
        super().__init__(*args, **kwargs)
        
        from django.db.models import Q
        from campaign_management.models import Campaign
        
        # Initialize an empty queryset
        collaterals = CMCollateral.objects.none()
        
        # Get the campaign if brand_campaign_id is provided
        campaign = None
        if brand_campaign_id:
            try:
                campaign = Campaign.objects.get(brand_campaign_id=brand_campaign_id)
                self.fields['campaign'].initial = campaign.brand_campaign_id
            except Campaign.DoesNotExist:
                pass
        
        # If we have an existing instance, get the campaign from it
        if self.instance and self.instance.pk:
            campaign = self.instance.campaign
            self.fields['campaign'].initial = campaign.brand_campaign_id
        
        # If we have a campaign, filter collaterals for that campaign
        if campaign:
            # Get collaterals from direct relationship (campaign field on Collateral)
            direct_collaterals = CMCollateral.objects.filter(campaign=campaign)
            
            # Get collaterals from many-to-many relationship (through CampaignCollateral)
            m2m_collaterals = CMCollateral.objects.filter(campaign_collaterals__campaign=campaign)
            
            # Combine both querysets
            collaterals = (direct_collaterals | m2m_collaterals).distinct()
            
            # If we're editing an existing instance, make sure to include the current collateral
            if self.instance and self.instance.pk:
                current_collateral = self.instance.collateral
                if current_collateral and not collaterals.filter(pk=current_collateral.pk).exists():
                    collaterals = collaterals | CMCollateral.objects.filter(pk=current_collateral.pk)
        else:
            # If no campaign is specified, show all collaterals
            collaterals = CMCollateral.objects.all()
        
        # Set the filtered queryset
        self.fields['collateral'].queryset = collaterals.order_by('title')
        
        # If we have an instance, set the initial value for the collateral field
        if self.instance and self.instance.pk and self.instance.collateral:
            self.fields['collateral'].initial = self.instance.collateral

    
    def clean(self):
        cleaned_data = super().clean()
        start_date = cleaned_data.get('start_date')
        end_date = cleaned_data.get('end_date')
        
        # Validate dates
        if start_date and end_date and end_date < start_date:
            raise ValidationError("End date cannot be earlier than start date.")
            
        # Keep campaign field for new records, but validate it exists
        if 'campaign' in cleaned_data and cleaned_data['campaign']:
            # Validate that the campaign exists
            from campaign_management.models import Campaign
            try:
                Campaign.objects.get(brand_campaign_id=cleaned_data['campaign'])
            except Campaign.DoesNotExist:
                raise ValidationError(f"Campaign with Brand Campaign ID '{cleaned_data['campaign']}' not found.")
        elif not self.instance.pk:
            # For new records, campaign is required
            if 'campaign' not in cleaned_data or not cleaned_data.get('campaign'):
                raise ValidationError("Brand Campaign ID is required for new campaign collateral.")
            
        return cleaned_data
    
    def save(self, commit=True):
        instance = super().save(commit=False)
        
        # For existing instances, preserve the original campaign
        if self.instance and self.instance.pk:
            original = CMCampaignCollateral.objects.get(pk=self.instance.pk)
            instance.campaign = original.campaign
        else:
            # For new instances, set campaign from the form data
            campaign_brand_id = self.cleaned_data.get('campaign')
            if campaign_brand_id:
                from campaign_management.models import Campaign
                try:
                    campaign = Campaign.objects.get(brand_campaign_id=campaign_brand_id)
                    instance.campaign = campaign
                except Campaign.DoesNotExist:
                    # This should have been caught in clean(), but handle gracefully
                    if commit:
                        raise ValidationError(f"Campaign with Brand Campaign ID '{campaign_brand_id}' not found.")

        # Normalize start/end into DateTime for the collateral_management.CampaignCollateral model
        # If date objects are provided (from <input type="date">), convert to start-of-day and end-of-day
        start_date_val = self.cleaned_data.get('start_date')
        end_date_val = self.cleaned_data.get('end_date')
        if start_date_val and not hasattr(start_date_val, 'hour'):
            instance.start_date = timezone.make_aware(datetime.combine(start_date_val, time.min))
        if end_date_val and not hasattr(end_date_val, 'hour'):
            # Use end of the day, drop microseconds for consistency
            instance.end_date = timezone.make_aware(datetime.combine(end_date_val, time.max.replace(microsecond=0)))
        
        if commit:
            instance.save()
        return instance


# Campaign Search Form
class CampaignSearchForm(forms.Form):
    brand_campaign_id = forms.CharField(
        required=False,
        label="Brand Campaign ID",
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Search by Brand Campaign ID'})
    )
    name = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Search by Campaign Name'})
    )
    brand_name = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Search by Brand Name'})
    )
    status = forms.ChoiceField(required=False, label="Status")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        from .models import Campaign
        # Try legacy constant first; otherwise read from the field
        choices = getattr(Campaign, "STATUS_CHOICES", None)
        if not choices:
            try:
                choices = Campaign._meta.get_field("status").choices
            except Exception:
                choices = ()
        self.fields["status"].choices = [("", "All Statuses")] + list(choices)


# Campaign Filter Form for Reports
class CampaignFilterForm(forms.Form):
    start_date = forms.DateField(
        required=False,
        label="From Date",
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
    )
    end_date = forms.DateField(
        required=False,
        label="To Date",
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
    )
    status = forms.ChoiceField(required=False, label="Status")
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        from .models import Campaign
        # Try legacy constant first; otherwise read from the field
        choices = getattr(Campaign, "STATUS_CHOICES", None)
        if not choices:
            try:
                choices = Campaign._meta.get_field("status").choices
            except Exception:
                choices = ()
        self.fields["status"].choices = [("", "All Statuses")] + list(choices)
    
    def clean(self):
        cleaned_data = super().clean()
        start_date = cleaned_data.get('start_date')
        end_date = cleaned_data.get('end_date')
        
        if start_date and end_date and end_date < start_date:
            raise ValidationError("End date cannot be earlier than start date.")
            
        return cleaned_data